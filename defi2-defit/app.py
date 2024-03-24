from openpyxl import load_workbook 
import networkx as nx
import geopy.distance
from flask import Flask, render_template, request

 
# Create a flask application
app = Flask(__name__)


@app.get('/')
def index():
    return render_template('index.html')

@app.get('/upload')
def upload():
    
    wb = load_workbook(filename='Cordonnees_GPS.xlsx', read_only=True)

    capitales_sheet = wb['Capitales_Wilaya']
    mougataas_sheet = wb['Mougataa']

    G = nx.Graph()

    for row in capitales_sheet.iter_rows(min_row=2, values_only=True):
        ville, latitude, longitude = row
        G.add_node(ville, pos=(latitude, longitude))

    for row in mougataas_sheet.iter_rows(min_row=2, values_only=True):
        ville,wilaya, latitude, longitude = row
        G.add_node(ville, pos=(latitude, longitude))

    for u in G.nodes():
        for v in G.nodes():
            if u != v and not G.has_edge(u, v):
                coords_u = G.nodes[u]['pos']
                coords_v = G.nodes[v]['pos']
                distance = geopy.distance.geodesic(coords_u, coords_v).km      
                G.add_edge(u, v, weight=distance)

    
    print("Nombre de nœuds :", G.number_of_nodes())
    print("Nombre d'arêtes :", G.number_of_edges())

    return render_template('upload.html')

@app.get('/maps')
def maps():
    return render_template('map.html')
