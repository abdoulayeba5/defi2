from openpyxl import load_workbook 
import networkx as nx
import geopy.distance

wb = load_workbook(filename='Cordonnees_GPS.xlsx', read_only=True)

capitales_sheet = wb['Capitales_Wilaya']
mougataas_sheet = wb['Mougataa']

G = nx.Graph()

for row in capitales_sheet.iter_rows(min_row=2, values_only=True):
    ville, latitude, longitude = row
    G.add_node(ville, pos=(latitude, longitude))

for row in mougataas_sheet.iter_rows(min_row=2, values_only=True):
    ville,wilaya, latitude, longitude = row
    G.add_node(ville, pos=(latitude, longitude))

for u in G.nodes():
    for v in G.nodes():
        if u != v and not G.has_edge(u, v):
            coords_u = G.nodes[u]['pos']
            coords_v = G.nodes[v]['pos']
            distance = geopy.distance.geodesic(coords_u, coords_v).km      
            G.add_edge(u, v, weight=distance)

 
print("Nombre de nœuds :", G.number_of_nodes())
print("Nombre d'arêtes :", G.number_of_edges())







# def plus_proche_voisin(G, depart):
#     chemin = [depart]
#     villes_non_visitees = set(G.nodes())
#     villes_non_visitees.remove(depart)

#     while villes_non_visitees:
#         ville_actuelle = chemin[-1]
#         voisins = G[ville_actuelle]
#         ville_suivante = min(voisins, key=lambda x: G[ville_actuelle][x]['weight'] if x in villes_non_visitees else float('inf'))
#         chemin.append(ville_suivante)
#         villes_non_visitees.remove(ville_suivante)

#     chemin.append(depart)  # Retour à la ville de départ
#     return chemin

# # Choix arbitraire de la ville de départ
# ville_depart = 'Nouakchott'

# # Application de l'algorithme de plus proche voisin
# chemin_approximation = plus_proche_voisin(G, ville_depart)

# # Affichage du chemin trouvé
# print("Chemin trouvé par l'algorithme de plus proche voisin :", chemin_approximation)




print("----------------------------------------------break---------------------------------------------")
import random
import numpy as np

def heuristique_fourmis(G, nb_fourmis, alpha=1, beta=2, evaporation_rate=0.5, iterations=1000):
    # Initialisation des phéromones sur toutes les arêtes
    pheromones = {(u, v): 1 for u, v in G.edges()}

    meilleur_chemin = None
    meilleure_distance = float('inf')

    for it in range(iterations):
        for k in range(nb_fourmis):
            ville_depart = random.choice(list(G.nodes()))
            chemin = [ville_depart]
            villes_restantes = set(G.nodes())
            villes_restantes.remove(ville_depart)

            while villes_restantes:
                ville_actuelle = chemin[-1]
                probabilites = []
                for voisin in villes_restantes:
                    probabilite = pheromones[(ville_actuelle, voisin)] ** alpha * (1 / G[ville_actuelle][voisin]['weight']) ** beta
                    probabilites.append((voisin, probabilite))
                somme_probabilites = sum(prob[1] for prob in probabilites)
                probabilites = [(voisin, prob / somme_probabilites) for voisin, prob in probabilites]
                prochain_ville = np.random.choice([v for v, _ in probabilites], p=[p for _, p in probabilites])
                chemin.append(prochain_ville)
                villes_restantes.remove(prochain_ville)

            chemin.append(ville_depart)  # Retour à la ville de départ

            # Évaluation du chemin
            distance_chemin = sum(G[chemin[i]][chemin[i+1]]['weight'] for i in range(len(chemin)-1))

            # Mise à jour du meilleur chemin trouvé
            if distance_chemin < meilleure_distance:
                meilleur_chemin = chemin
                meilleure_distance = distance_chemin

            # Mise à jour des phéromones sur les arêtes du chemin
            for i in range(len(chemin)-1):
                pheromones[(chemin[i], chemin[i+1])] += 1 / distance_chemin

        # Évaporation des phéromones sur toutes les arêtes
        for edge in pheromones:
            pheromones[edge] *= (1 - evaporation_rate)

    return meilleur_chemin, meilleure_distance

# Paramètres de l'heuristique des fourmis
nb_fourmis = 10
alpha = 1  # Importance des phéromones
beta = 2  # Importance de la distance
evaporation_rate = 0.5  # Taux d'évaporation des phéromones
iterations = 100  # Nombre d'itérations

# Application de l'heuristique des fourmis
meilleur_chemin_fourmis, meilleure_distance_fourmis = heuristique_fourmis(G, nb_fourmis, alpha, beta, evaporation_rate, iterations)

# Affichage du meilleur chemin trouvé par l'heuristique des fourmis
print("Meilleur chemin trouvé par l'heuristique des fourmis :", meilleur_chemin_fourmis)
print("Longueur du chemin :", meilleure_distance_fourmis)



























































# Affichage de tous les nœuds avec leurs coordonnées GPS
# print("Liste de tous les nœuds avec leurs coordonnées GPS :")
# for node in G.nodes(data=True):
#     ville = node[0]
#     coords = node[1]['pos']
#     print(f"{ville} : {coords}")








# print("Distances entre les villes :")
# for u, v, data in G.edges(data=True):
#     distance = data['weight']
#     print(f"Distance entre {u} et {v} : {distance} km")
    











# from openpyxl import Workbook

# # Création d'un nouveau classeur Excel
# wb_new = Workbook()
# ws = wb_new.active

# # Entêtes de colonnes
# ws.append(["Ville 1", "Ville 2", "Distance (km)"])

# # Écriture des données
# for u, v, data in G.edges(data=True):
#     distance = data['weight']
#     ws.append([u, v, distance])

# # Enregistrement du fichier Excel
# wb_new.save("distances_entre_villes.xlsx")
